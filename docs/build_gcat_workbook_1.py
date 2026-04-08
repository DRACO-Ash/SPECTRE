#!/usr/bin/env python3
"""
GCAT Comprehensive Excel Workbook Builder v2.0
================================================
Downloads ALL public TSV datasets from Jonathan McDowell's General Catalog
of Artificial Space Objects and builds a comprehensive, filterable Excel workbook.

Data source: https://planet4589.org/space/gcat/ (CC-BY licence)
Citation: data from GCAT (J. McDowell, planet4589.org/space/gcat)

Requirements: pip install pandas openpyxl
Usage:        python build_gcat_workbook.py
Output:       GCAT_Workbook.xlsx

Datasets (30+):
  Supporting Tables:  Organisations, Launch Sites, Launch Points, Platforms,
                      LV Families, Launch Vehicles, LV Stages, Engines,
                      References, Worlds
  Object Catalogs:    SatCat, AuxCat, FtoCat, EventCat, DeepCat, HelioCat,
                      LunarPlanetCat, LanderCat, TmpCat
  Payload Catalogs:   PayloadSatCat, PayloadAuxCat, PayloadFtoCat, PayloadDeepCat
  Derived Catalogs:   CurrentCat, LaunchLog, GeosyncCat, ActiveCat
  Full Launch List:   Complete launch data
  Analysis:           Yearly Summary, By Vehicle, By Site, By State
"""

import urllib.request
import pandas as pd
import io
import re
import time
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = 'https://planet4589.org/space/gcat'

# Ordered dict: tab name -> URL.  Order here = tab order in workbook.
DATASETS = {
    # ── SUPPORTING TABLES ──
    'Organisations':     f'{BASE}/tsv/tables/orgs.tsv',
    'Launch Sites':      f'{BASE}/tsv/tables/sites.tsv',
    'Launch Points':     f'{BASE}/tsv/tables/lp.tsv',
    'Platforms':         f'{BASE}/tsv/tables/platforms.tsv',
    'LV Families':       f'{BASE}/tsv/tables/family.tsv',
    'Launch Vehicles':   f'{BASE}/tsv/tables/lv.tsv',
    'LV Stages':         f'{BASE}/tsv/tables/stages.tsv',
    'Engines':           f'{BASE}/tsv/tables/engines.tsv',
    'References':        f'{BASE}/tsv/tables/refs.tsv',
    'Worlds':            f'{BASE}/tsv/worlds/worlds.tsv',
    # ── DERIVED CATALOGS (most operationally useful first) ──
    'Current Catalog':   f'{BASE}/tsv/derived/currentcat.tsv',
    'Launch Log':        f'{BASE}/tsv/derived/launchlog.tsv',
    'Active Catalog':    f'{BASE}/tsv/derived/active.tsv',
    'Geosync Catalog':   f'{BASE}/tsv/derived/geotab.tsv',
    # ── FULL LAUNCH LIST ──
    'Full Launches':     f'{BASE}/tsv/launch/launch.tsv',
    # ── OBJECT CATALOGS ──
    'SatCat':            f'{BASE}/tsv/cat/satcat.tsv',
    'AuxCat':            f'{BASE}/tsv/cat/auxcat.tsv',
    'FtoCat':            f'{BASE}/tsv/cat/ftocat.tsv',
    'EventCat':          f'{BASE}/tsv/cat/ecat.tsv',
    'DeepCat':           f'{BASE}/tsv/cat/deepcat.tsv',
    'HelioCat':          f'{BASE}/tsv/cat/hcocat.tsv',
    'LunarPlanetCat':    f'{BASE}/tsv/cat/lprcat.tsv',
    'LanderCat':         f'{BASE}/tsv/cat/landercat.tsv',
    'TmpCat':            f'{BASE}/tsv/cat/tmpcat.tsv',
    # ── PAYLOAD CATALOGS ──
    'PayloadSatCat':     f'{BASE}/tsv/cat/psatcat.tsv',
    'PayloadAuxCat':     f'{BASE}/tsv/cat/pauxcat.tsv',
    'PayloadFtoCat':     f'{BASE}/tsv/cat/pftocat.tsv',
    'PayloadDeepCat':    f'{BASE}/tsv/cat/pdeepcat.tsv',
}

HEADER_FILL = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Arial', size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))

TAB_COLOURS = {
    'README': '1B3A5C',
    'Organisations': '27AE60', 'Launch Sites': '2ECC71', 'Launch Points': '1ABC9C',
    'Platforms': '16A085', 'LV Families': '1B8A6B', 'Launch Vehicles': '138D75',
    'LV Stages': '0E6655', 'Engines': '0B5345', 'References': '117A65', 'Worlds': '148F77',
    'SatCat': '2980B9', 'AuxCat': '3498DB', 'FtoCat': '2471A3', 'EventCat': '1F618D',
    'DeepCat': '1A5276', 'HelioCat': '154360', 'LunarPlanetCat': '1B4F72',
    'LanderCat': '21618C', 'TmpCat': '2E86C1',
    'PayloadSatCat': 'E67E22', 'PayloadAuxCat': 'F39C12',
    'PayloadFtoCat': 'D68910', 'PayloadDeepCat': 'CA6F1E',
    'Current Catalog': 'C0392B', 'Launch Log': 'E74C3C',
    'Geosync Catalog': '8E44AD', 'Active Catalog': '9B59B6', 'Full Launches': '7D3C98',
    'Yearly Summary': 'C0392B', 'By Vehicle': '16A085',
    'By Site': 'F39C12', 'By State': '2C3E50',
}


def download_tsv(url, name):
    print(f'  {name:.<30s} ', end='', flush=True)
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'GCAT-Workbook/2.0'})
        with urllib.request.urlopen(req, timeout=180) as resp:
            raw = resp.read().decode('utf-8', errors='replace')
    except Exception as e:
        print(f'FAILED ({e})')
        return None
    lines = [l for l in raw.split('\n') if l.strip() and not l.startswith('# ')]
    if not lines:
        print('EMPTY')
        return None
    header = lines[0].lstrip('#').strip()
    tsv = header + '\n' + '\n'.join(lines[1:])
    try:
        df = pd.read_csv(io.StringIO(tsv), sep='\t', dtype=str,
                         on_bad_lines='skip', low_memory=False)
    except Exception as e:
        print(f'PARSE ERROR ({e})')
        return None
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].apply(lambda x: x.strip() if isinstance(x, str) else x)
    df = df.replace(['-', '*'], pd.NA)
    print(f'{len(df):>8,} rows x {len(df.columns):>3} cols')
    return df


def style_sheet(ws, ncols, nrows, sr=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=sr, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lim = min(400, nrows)
    for r in range(sr + 1, sr + lim + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
    for c in range(1, ncols + 1):
        mx = len(str(ws.cell(row=sr, column=c).value or ''))
        for r in range(sr + 1, min(sr + 60, sr + nrows + 1)):
            v = ws.cell(row=r, column=c).value
            if v:
                mx = max(mx, min(len(str(v)), 35))
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 3, 40)


def add_table(ws, ncols, nrows, name, sr=1):
    if nrows == 0:
        return
    ref = f'A{sr}:{get_column_letter(ncols)}{sr + nrows}'
    safe = re.sub(r'[^A-Za-z0-9_]', '_', name)[:50]
    tbl = Table(displayName=safe, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


def write_df(ws, df, sr=1):
    for ci, cn in enumerate(df.columns, 1):
        ws.cell(row=sr, column=ci, value=cn)
    for ri, (_, row) in enumerate(df.iterrows(), sr + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value='' if pd.isna(val) else val)


def build_sheet(wb, name, df, colour=None):
    if df is None or df.empty:
        return
    tab = name[:31]
    ws = wb.create_sheet(tab)
    if colour:
        ws.sheet_properties.tabColor = colour
    write_df(ws, df)
    style_sheet(ws, len(df.columns), len(df))
    add_table(ws, len(df.columns), len(df), name)
    ws.freeze_panes = 'A2'
    return ws


def parse_date(s):
    if pd.isna(s):
        return pd.NaT
    s = re.sub(r'\?$', '', str(s).strip())
    s = re.sub(r'\s+', ' ', s).strip()
    for fmt in ['%Y %b %d %H%M:%S', '%Y %b %d %H%M', '%Y %b %d']:
        try:
            return pd.to_datetime(s, format=fmt)
        except Exception:
            continue
    try:
        return pd.to_datetime(s[:11].strip(), format='%Y %b %d')
    except Exception:
        return pd.NaT


def main():
    ts = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
    print('=' * 70)
    print('  GCAT Comprehensive Workbook Builder v2.0')
    print(f'  {ts}')
    print('=' * 70)

    # Download
    print('\n  DOWNLOADING...\n')
    data = {}
    for name, url in DATASETS.items():
        data[name] = download_tsv(url, name)
        time.sleep(0.2)

    # Summaries
    print('\n  COMPUTING SUMMARIES...')
    summaries = {}
    ll = data.get('Launch Log')
    if ll is not None and not ll.empty:
        ll['_Date'] = ll['Launch_Date'].apply(parse_date)
        ll['Year'] = ll['_Date'].dt.year
        ll['LaunchOutcome'] = ll.get('Launch_Code', pd.Series(dtype=str)).map(
            lambda x: {'OS': 'Success', 'OF': 'Failure', 'DS': 'Deep Space',
                        'OS75': 'Partial', 'OF40': 'Partial Orbit'
                        }.get(str(x).strip(), str(x)) if pd.notna(x) else '')
        dedup = ll.drop_duplicates(subset=['Launch_Tag'], keep='first')

        y = dedup.groupby(['Year', 'LVState']).size().reset_index(name='Launches')
        y = y[y['Year'].notna()].copy()
        y['Year'] = y['Year'].astype(int)
        yp = y.pivot_table(index='Year', columns='LVState', values='Launches',
                           fill_value=0, aggfunc='sum')
        yp['Total'] = yp.sum(axis=1)
        summaries['Yearly Summary'] = yp.reset_index()

        lv = dedup['LV_Type'].value_counts().head(50).reset_index()
        lv.columns = ['Launch_Vehicle', 'Total_Launches']
        summaries['By Vehicle'] = lv

        si = dedup['Launch_Site'].value_counts().head(50).reset_index()
        si.columns = ['Launch_Site', 'Total_Launches']
        summaries['By Site'] = si

    cc = data.get('Current Catalog')
    if cc is not None and not cc.empty and 'State' in cc.columns:
        ap = cc[(cc['Type'].str.contains('P', na=False)) &
                (cc['ExpandedStatus'].str.contains('orbit', case=False, na=False))]
        sc = ap['State'].value_counts().head(50).reset_index()
        sc.columns = ['State', 'Active_Payloads']
        summaries['By State'] = sc

    # Write workbook
    print('\n  WRITING WORKBOOK...\n')
    wb = Workbook()
    ws = wb.active
    ws.title = 'README'
    ws.sheet_properties.tabColor = '1B3A5C'
    ws['A1'] = 'GCAT Comprehensive Workbook'
    ws['A1'].font = Font(name='Arial', bold=True, size=20, color='1B3A5C')
    ws['A2'] = 'General Catalog of Artificial Space Objects'
    ws['A2'].font = Font(name='Arial', size=11, color='666666')
    ws['A3'] = 'Jonathan C. McDowell — planet4589.org/space/gcat'
    ws['A3'].font = Font(name='Arial', size=10, color='888888')
    ws['A5'] = f'Generated: {ts}'
    ws['A5'].font = DATA_FONT
    ws['A6'] = 'Licence: Creative Commons CC-BY'
    ws['A6'].font = DATA_FONT

    r = 8
    sections = [
        ('SUPPORTING TABLES', [
            ('Organisations', 'Countries, agencies, companies, owners, manufacturers'),
            ('Launch Sites', 'Launch origins / cosmodromes with coordinates'),
            ('Launch Points', 'Individual pads / launch positions within sites'),
            ('Platforms', 'Mobile launch platforms (ships, aircraft, mobile pads)'),
            ('LV Families', 'Launch vehicle family groupings'),
            ('Launch Vehicles', 'Launch vehicle type definitions'),
            ('LV Stages', 'Rocket stage specifications'),
            ('Engines', 'Rocket engine catalog'),
            ('References', 'Launch time citation sources'),
            ('Worlds', 'Solar system bodies (central body definitions)'),
        ]),
        ('DERIVED CATALOGS', [
            ('Current Catalog', 'Current status of every tracked object'),
            ('Launch Log', 'Orbital launch log with payloads'),
            ('Active Catalog', 'Currently operational payloads'),
            ('Geosync Catalog', 'Geosynchronous / geostationary objects'),
            ('Full Launches', 'Complete launch data by vehicle family'),
        ]),
        ('OBJECT CATALOGS', [
            ('SatCat', 'Standard satellite catalog (US Space Force tracked)'),
            ('AuxCat', 'Auxiliary objects not in US catalog'),
            ('FtoCat', 'Objects from failed launches'),
            ('EventCat', 'Object phase changes (dockings, manoeuvres, reentries)'),
            ('DeepCat', 'Deep space objects'),
            ('HelioCat', 'Heliocentric orbit register'),
            ('LunarPlanetCat', 'Lunar and planetary orbit register'),
            ('LanderCat', 'Lunar and planetary landings and impacts'),
            ('TmpCat', 'Temporary catalog (awaiting permanent assignment)'),
        ]),
        ('PAYLOAD CATALOGS', [
            ('PayloadSatCat', 'Payload metadata: category, end-of-life, civil/military/commercial'),
            ('PayloadAuxCat', 'Payload metadata for auxiliary catalog objects'),
            ('PayloadFtoCat', 'Payload metadata for failed-to-orbit objects'),
            ('PayloadDeepCat', 'Payload metadata for deep space objects'),
        ]),
        ('ANALYSIS SUMMARIES', [
            ('Yearly Summary', 'Launches per year by launching state'),
            ('By Vehicle', 'Top 50 launch vehicles by total launches'),
            ('By Site', 'Top 50 launch sites by total launches'),
            ('By State', 'Top 50 states by active payloads in orbit'),
        ]),
    ]

    for section, tabs in sections:
        ws.cell(row=r, column=1, value=section).font = Font(
            name='Arial', bold=True, size=11, color='1B3A5C')
        r += 1
        for tab, desc in tabs:
            df = data.get(tab) or summaries.get(tab)
            cnt = f'{len(df):,} rows' if df is not None and not df.empty else 'N/A'
            ws.cell(row=r, column=1, value=tab[:31]).font = Font(
                name='Arial', bold=True, size=9, color='333333')
            ws.cell(row=r, column=2, value=desc).font = DATA_FONT
            ws.cell(row=r, column=3, value=cnt).font = Font(
                name='Arial', size=9, color='888888')
            r += 1
        r += 1

    ws.cell(row=r, column=1, value='TO REFRESH:').font = Font(
        name='Arial', bold=True, size=10, color='C0392B')
    ws.cell(row=r, column=2, value='python build_gcat_workbook.py').font = Font(
        name='Arial', size=10, color='C0392B')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 16

    # Write all data tabs
    for name in DATASETS:
        df = data.get(name)
        if df is not None and not df.empty:
            build_sheet(wb, name, df, TAB_COLOURS.get(name))
            print(f'  + {name[:30]:.<32s} {len(df):>8,} rows')

    for name, df in summaries.items():
        if df is not None and not df.empty:
            build_sheet(wb, name, df, TAB_COLOURS.get(name))
            print(f'  + {name[:30]:.<32s} {len(df):>8,} rows')

    output = 'GCAT_Workbook.xlsx'
    wb.save(output)
    print(f'\n{"=" * 70}')
    print(f'  DONE -> {output}')
    print(f'  Sheets: {len(wb.sheetnames)}')
    print(f'{"=" * 70}')


if __name__ == '__main__':
    main()
